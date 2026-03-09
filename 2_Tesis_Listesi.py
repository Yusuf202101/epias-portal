import streamlit as st
import requests
import pandas as pd
import json
import time
import io
from datetime import datetime, timedelta, date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Tesis Listesi", page_icon="🏭", layout="wide")
st.title("🏭 Organizasyon & Tesis Listesi")
st.caption("EPİAŞ Şeffaflık Platformu — UEVCB Tesis Sorgulama")

URL_AUTH     = "https://giris.epias.com.tr/cas/v1/tickets"
URL_ORG_LIST = "https://seffaflik.epias.com.tr/electricity-service/v1/generation/data/organization-list"
URL_UEVCB    = "https://seffaflik.epias.com.tr/electricity-service/v1/generation/data/uevcb-list"

def get_tgt(username, password):
    now = datetime.now()
    if "tgt" in st.session_state and "tgt_time" in st.session_state:
        if now - st.session_state["tgt_time"] < timedelta(hours=2):
            return st.session_state["tgt"]
    r = requests.post(URL_AUTH,
        data={"username": username, "password": password},
        headers={"Content-Type": "application/x-www-form-urlencoded"})
    if r.status_code == 201:
        tgt = r.headers["Location"].split("/")[-1]
        st.session_state["tgt"] = tgt
        st.session_state["tgt_time"] = now
        return tgt
    st.error(f"❌ TGT alınamadı. HTTP {r.status_code}")
    return None

def excel_olustur(sonuc_listesi, start_date, end_date):
    satirlar   = []
    tesis_keys = []
    for kayit in sonuc_listesi:
        org_id   = kayit.get("organizationId", "")
        org_name = kayit.get("organizationName", "")
        tesisler = kayit.get("uevcbListesi", [])
        if not tesisler:
            satirlar.append({"Organizasyon ID": org_id, "Organizasyon Adı": org_name})
        else:
            for tesis in tesisler:
                satir    = {"Organizasyon ID": org_id, "Organizasyon Adı": org_name}
                alan_map = {"id": "Tesis ID", "name": "Tesis Adı", "eic": "EIC"}
                for k, v in tesis.items():
                    temiz_k = alan_map.get(k, k)
                    satir[temiz_k] = v
                    if temiz_k not in tesis_keys:
                        tesis_keys.append(temiz_k)
                satirlar.append(satir)

    tum_kolonlar = ["Organizasyon ID", "Organizasyon Adı"] + tesis_keys
    HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    HEADER_FILL  = PatternFill("solid", start_color="1F4E79")
    DATA_FONT    = Font(name="Arial", size=9)
    ALT_FILL     = PatternFill("solid", start_color="EBF3FB")
    CENTER       = Alignment(horizontal="center", vertical="center")
    LEFT         = Alignment(horizontal="left",   vertical="center")
    thin         = Side(style="thin", color="BDD7EE")
    BORDER       = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    ws = wb.active
    ws.title = "Tüm Tesisler"
    ws.merge_cells(f"A1:{get_column_letter(len(tum_kolonlar))}1")
    ws["A1"] = f"ORGANİZASYON & TESİS LİSTESİ  |  {str(start_date)[:10]} – {str(end_date)[:10]}  |  Toplam {len(satirlar)} kayıt"
    ws["A1"].font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", start_color="1F4E79")
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

    for col, baslik in enumerate(tum_kolonlar, 1):
        cell = ws.cell(row=2, column=col, value=baslik)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL
        cell.alignment = CENTER; cell.border = BORDER
    ws.row_dimensions[2].height = 20

    for i, satir in enumerate(satirlar, 1):
        row  = i + 2
        fill = ALT_FILL if i % 2 == 0 else None
        for col, kolon in enumerate(tum_kolonlar, 1):
            cell = ws.cell(row=row, column=col, value=satir.get(kolon, ""))
            cell.font = DATA_FONT
            cell.alignment = LEFT if col == 2 else CENTER
            cell.border = BORDER
            if fill: cell.fill = fill

    col_widths = {"Organizasyon ID": 18, "Organizasyon Adı": 42}
    for col, kolon in enumerate(tum_kolonlar, 1):
        w = col_widths.get(kolon) or min(max(len(str(kolon)), max((len(str(s.get(kolon,""))) for s in satirlar), default=0)) + 4, 35)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ── SIDEBAR ──────────────────────────────────
with st.sidebar:
    st.header("🔐 Giriş Bilgileri")
    default_user = st.secrets.get("epias", {}).get("username", "") if hasattr(st, "secrets") else ""
    default_pass = st.secrets.get("epias", {}).get("password", "") if hasattr(st, "secrets") else ""
    username = st.text_input("EPİAŞ Kullanıcı Adı", value=default_user)
    password = st.text_input("EPİAŞ Şifre", value=default_pass, type="password")
    st.divider()
    st.header("📅 Tarih Aralığı")
    bugun      = date.today()
    start_date = st.date_input("Başlangıç", value=bugun.replace(day=1))
    end_date   = st.date_input("Bitiş",     value=bugun)
    st.divider()
    baslat = st.button("🚀 Sorguyu Başlat", use_container_width=True, type="primary")

# ── ANA İŞLEM ────────────────────────────────
if baslat:
    if not username or not password:
        st.warning("⚠️ Kullanıcı adı ve şifre gerekli.")
        st.stop()
    if start_date > end_date:
        st.error("❌ Başlangıç tarihi bitiş tarihinden büyük olamaz.")
        st.stop()

    start_iso = f"{start_date}T00:00:00+03:00"
    end_iso   = f"{end_date}T00:00:00+03:00"

    with st.spinner("🔐 Yetkilendiriliyor..."):
        tgt = get_tgt(username, password)
    if not tgt:
        st.stop()

    with st.spinner("📋 Organizasyon listesi çekiliyor..."):
        try:
            r = requests.post(URL_ORG_LIST, headers={"TGT": tgt, "Content-Type": "application/json"},
                              json={"startDate": start_iso, "endDate": end_iso})
            r.raise_for_status()
            org_items = r.json().get("items", [])
        except Exception as e:
            st.error(f"❌ Organizasyon listesi alınamadı: {e}")
            st.stop()

    toplam_org = len(org_items)
    st.info(f"📋 {toplam_org} organizasyon bulundu. Tesisler çekiliyor...")

    sonuc_listesi = []
    progress_bar  = st.progress(0)
    durum_text    = st.empty()

    for index, org in enumerate(org_items, 1):
        org_id   = org.get("organizationId") or org.get("id")
        org_name = org.get("organizationName") or org.get("name")
        durum_text.text(f"[{index}/{toplam_org}] {org_name}")
        progress_bar.progress(index / toplam_org)
        try:
            r = requests.post(URL_UEVCB,
                headers={"TGT": tgt, "Content-Type": "application/json"},
                json={"organizationId": str(org_id), "startDate": start_iso, "endDate": end_iso})
            tesisler = r.json().get("items", []) if r.status_code == 200 else []
        except Exception:
            tesisler = []
        sonuc_listesi.append({"organizationId": org_id, "organizationName": org_name, "uevcbListesi": tesisler})
        time.sleep(0.3)

    durum_text.empty()
    progress_bar.empty()

    toplam_tesis = sum(len(k["uevcbListesi"]) for k in sonuc_listesi)
    bos_org      = sum(1 for k in sonuc_listesi if not k["uevcbListesi"])
    m1, m2, m3   = st.columns(3)
    m1.metric("Toplam Organizasyon", toplam_org)
    m2.metric("Toplam Tesis",        toplam_tesis)
    m3.metric("Tesisi Olmayan Org.", bos_org)
    st.divider()

    satirlar = []
    for kayit in sonuc_listesi:
        oid, oname = kayit["organizationId"], kayit["organizationName"]
        if not kayit["uevcbListesi"]:
            satirlar.append({"Organizasyon ID": oid, "Organizasyon Adı": oname})
        else:
            alan_map = {"id": "Tesis ID", "name": "Tesis Adı", "eic": "EIC"}
            for t in kayit["uevcbListesi"]:
                row = {"Organizasyon ID": oid, "Organizasyon Adı": oname}
                for k, v in t.items():
                    row[alan_map.get(k, k)] = v
                satirlar.append(row)

    df = pd.DataFrame(satirlar)
    st.subheader(f"📊 Önizleme — {len(df)} satır")
    st.dataframe(df, use_container_width=True, height=400)
    st.divider()

    col_dl1, col_dl2 = st.columns(2)
    with col_dl1:
        st.download_button("📥 Excel İndir (.xlsx)",
            data=excel_olustur(sonuc_listesi, start_date, end_date),
            file_name=f"tesisler_{start_date}_{end_date}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)
    with col_dl2:
        st.download_button("📥 JSON İndir (.json)",
            data=json.dumps(sonuc_listesi, ensure_ascii=False, indent=4).encode("utf-8"),
            file_name=f"tesisler_{start_date}_{end_date}.json",
            mime="application/json",
            use_container_width=True)
